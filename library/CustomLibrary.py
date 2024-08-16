import openpyxl
from robot.libraries.BuiltIn import BuiltIn
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import requests
import win32com.client as win32
import os
from selenium.webdriver.common.action_chains import ActionChains
import pyautogui
import allure
import subprocess
from PyPDF2 import PdfReader
import json

class CustomLibrary(object):
        @property
        def _sel_lib(self):
            return BuiltIn().get_library_instance('SeleniumLibrary')

        @property
        def _driver(self):
            return self._sel_lib.driver

        def open_chrome_browser(self,url):
            """Return the True if Chrome browser opened """
            selenium = BuiltIn().get_library_instance('SeleniumLibrary')
            try:
                options = webdriver.ChromeOptions()
                options.add_argument('--disable-gpu')
                options.add_argument("disable-extensions")
                options.add_argument('--ignore-ssl-errors=yes')
                options.add_argument('--ignore-certificate-errors')
                options.add_argument('--use-fake-ui-for-media-stream')
                options.add_experimental_option('prefs', {
                    'credentials_enable_service': False,
                    'profile': {
                        'password_manager_enabled': False
                    }
                })
                options.add_experimental_option("excludeSwitches",["enable-automation","load-extension"])
                selenium.create_webdriver('Chrome',chrome_options=options)
                selenium.go_to(url)
                return True
            except:
                return False
            
        def convert_xls_2_xlsx(self, xls_path, xlsx_path):
            # Create temp xlsx-File
            if os.path.exists(xlsx_path): os.remove(xlsx_path)
            excel = win32.Dispatch("Excel.Application")
            excel.Visible = 0
            wb = excel.Workbooks.Open(xls_path)
            wb.SaveAs(xlsx_path, FileFormat = 51)    #FileFormat = 51 is for .xlsx extension
            wb.Close()

        def create_dictionary_from_two_lists(self,key_list,value_list):
            # using dict() and zip() to convert lists to dictionary
            res = dict(zip(key_list, value_list))
            return  res

        def open_file(self, path):
            os.system(path)

        def print_screen(self):
            pyautogui.FAILSAFE = False
            pyautogui.keyDown("printscreen")
            pyautogui.keyUp("printscreen")
            time.sleep(2)

        def open_headless_chrome(self,url):
            """Return the True if Chrome browser opened """
            selenium = BuiltIn().get_library_instance('SeleniumLibrary')
            options = webdriver.ChromeOptions()
            options.add_argument("--window-size=1440,900")
            options.add_argument('--disable-gpu')
            options.add_argument("disable-extensions")
            options.add_argument("--headless")
            options.add_experimental_option('prefs', {
                'credentials_enable_service': False,
                'profile': {
                    'password_manager_enabled': False
                }
            })
            options.add_experimental_option("excludeSwitches",["enable-automation","load-extension"])
            selenium.create_webdriver('Chrome',chrome_options=options)
            selenium.go_to(url)
        
        def javascript_click(self, locator):
            try:
                element = self._sel_lib.get_webelement(locator)
                self._driver.execute_script("arguments[0].click();", element)
            except Exception as e:
                # If an exception occurs, take a screenshot
                filename = time.strftime("%H%M%S")
                self.screenshot_page(filename)
                raise AssertionError(f"Failed due to: {e}")
            
        def get_text_by_using_javascript(self, locator):
            try:
                element = self._sel_lib.get_webelement(locator) 
                return self._driver.execute_script("return arguments[0].textContent;", element)
            except Exception as e:
                # If an exception occurs, take a screenshot
                filename = time.strftime("%H%M%S")
                self.screenshot_page(filename)
                raise AssertionError(f"Failed due to: {e}")

        def wait_until_time(self,arg):
                time.sleep(int(arg))
            
        def wait_until_element_clickable(self,locator):
            try:
                """ An Expectation for checking that an element is either invisible or not present on the DOM."""
                if locator.startswith("//") or locator.startswith("(//"):
                    WebDriverWait(self._driver, 60).until(EC.element_to_be_clickable((By.XPATH, locator)))
                else:
                    WebDriverWait(self._driver, 60).until(EC.element_to_be_clickable((By.ID, locator)))
            except Exception as e:
                # If an exception occurs, take a screenshot
                filename = time.strftime("%H%M%S")
                self.screenshot_page(filename)
                raise AssertionError(f"Failed due to: {e}")
        
        def get_ms_excel_row_values_into_dictionary_based_on_key(self, filepath, keyName, sheetName=None):
            """Returns the dictionary of values given row in the MS Excel file"""
            workbook = openpyxl.load_workbook(filepath)
            snames = workbook.sheetnames
            dictVar = {}

            if sheetName is None:
                sheetName = snames[0]

            if sheetName not in snames or not self.Verify_the_sheet_in_ms_excel_file(filepath, sheetName):
                return dictVar

            worksheet = workbook[sheetName]
            headersList = [str(cell.value) for cell in worksheet[1]]

            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if str(row[0]) != str(keyName):
                    continue

                for rowIndex, cell_data in enumerate(row):
                    if cell_data is None or cell_data == "":
                        continue

                    # Ensure that get_unique_test_data is implemented correctly
                    cell_data = self.get_unique_test_data(cell_data)

                    dictVar[str(headersList[rowIndex])] = str(cell_data)

            return dictVar                 

        def get_unique_test_data(self,testdata):
            """Returns the unique if data contains unique word """
            ts = time.strftime("%H%M%S")
            unique_string = str(ts)
            testdata = testdata.replace("UNIQUE",unique_string)
            testdata = testdata.replace("Unique",unique_string)
            testdata = testdata.replace("unique",unique_string)
            return testdata

        def Verify_the_sheet_in_ms_excel_file(self,filepath,sheetName):
            """Returns the True if the specified work sheets exist in the specifed MS Excel file else False"""
            # workbook = xlrd.open_workbook(filepath)
            # snames = workbook.sheet_names()
            workbook = openpyxl.load_workbook(filepath)
            snames = workbook.sheetnames
            sStatus = False        
            if sheetName == None:
                return True
            else:
                for sname in snames:
                    if sname.lower() == sheetName.lower():
                        wsname = sname
                        sStatus = True
                        break
                if sStatus == False:
                    print ("Error: The specified sheet: "+str(sheetName)+" doesn't exist in the specified file: " +str(filepath))
            return sStatus
        
        def clear_text_field(self, locator):
            try:
                element = self._sel_lib.get_webelement(locator)
                self._driver.execute_script('arguments[0].value = "";', element)
            except Exception as e:
                filename = time.strftime("%H%M%S")
                self.screenshot_page(filename)
                raise AssertionError(f"Failed due to: {e}")

        def javascript_input_text(self,locator, text):
            try:
                element = self._sel_lib.get_webelement(locator)
                self._driver.execute_script('arguments[0].value = arguments[1];', element, text)
                self._driver.execute_script("arguments[0].dispatchEvent(new Event('input', { bubbles: true }));", element)
                self._driver.execute_script('arguments[0].focus();', element)
            except Exception as e:
                # If an exception occurs, take a screenshot
                filename = time.strftime("%H%M%S")
                self.screenshot_page(filename)
                raise AssertionError(f"Failed due to: {e}")

        def download_Pdf_opened_in_browser(self, project_folder, filename):
            pdf_url = self._driver.current_url
            response = requests.get(pdf_url)
            file_name = os.path.join(project_folder, filename)
            with open(file_name, 'wb') as f:
                f.write(response.content)

        def click_element_with_offset(self, locator, offset_x, offset_y):
            try:
                element = self._sel_lib.get_webelement(locator)
                ActionChains(self._driver).move_to_element_with_offset(element, offset_x, offset_y).click().perform()
            except Exception as e:
                # If an exception occurs, take a screenshot
                filename = time.strftime("%H%M%S")
                self.screenshot_page(filename)
                raise AssertionError(f"Failed due to: {e}")

        def right_click_element_with_offset(self, locator, x, y):
            try:
                element = self._sel_lib.get_webelement(locator)
                ActionChains(self._driver).move_to_element_with_offset(element, x, y).context_click().perform()
            except Exception as e:
                # If an exception occurs, take a screenshot
                filename = time.strftime("%H%M%S")
                self.screenshot_page(filename)
                raise AssertionError(f"Failed due to: {e}")
    
        def switch_to_parent_frame(self):
            self._driver.switch_to.parent_frame()

        def screenshot_page(self,png_name):
            ul = BuiltIn().get_library_instance('SeleniumLibrary')
            path = ul.capture_page_screenshot()
            allure.attach.file(path, name=png_name, attachment_type=allure.attachment_type.JPG)
            return path

        def open_file_and_take_screenshot(self, path, file_name, kill='None', app='None'):
            subprocess.Popen([path], shell=True)
            time.sleep(7)
            screenshot = pyautogui.screenshot()
            cur_time = time.strftime("%H%M%S")
            file = file_name + cur_time+'.png'
            screenshot.save(file)
            if kill!='None': os.system("taskkill /f /im "+ app +".exe")
                           
        def press_keyboard_key(self, key_name):
            try:
                pyautogui.press(key_name)
            except Exception as e:
                # If an exception occurs, take a screenshot
                filename = time.strftime("%H%M%S")
                self.screenshot_page(filename)
                raise AssertionError(f"Failed due to: {e}")
 
        def press_multiple_keyboard_keys(self, key_name1, key_name2, key_name3):
            try:
                pyautogui.hotkey(key_name1, key_name2, key_name3)
            except Exception as e:
                # If an exception occurs, take a screenshot
                filename = time.strftime("%H%M%S")
                self.screenshot_page(filename)
                raise AssertionError(f"Failed due to: {e}")
        
        def press_keyboard_keypairs(self, key_name1, key_name2):
            try:
                pyautogui.hotkey(key_name1, key_name2)
            except Exception as e:
                # If an exception occurs, take a screenshot
                filename = time.strftime("%H%M%S")
                self.screenshot_page(filename)
                raise AssertionError(f"Failed due to: {e}")

        def input_text_with_offset(self, locator, offset_x, offset_y, text):
            try:
                element = self._sel_lib.get_webelement(locator)
                ActionChains(self._driver).move_to_element_with_offset(element, offset_x, offset_y).click().send_keys(text).perform()
            except Exception as e:
                filename = time.strftime("%H%M%S")
                self.screenshot_page(filename)
                raise AssertionError(f"Failed due to: {e}")
        
        def double_click_element_with_offset(self, locator, offset_x, offset_y):
            try:
                element = self._sel_lib.get_webelement(locator)
                ActionChains(self._driver).move_to_element_with_offset(element, offset_x, offset_y).double_click().perform()
            except Exception as e:
                filename = time.strftime("%H%M%S")
                self.screenshot_page(filename)
                raise AssertionError(f"Failed due to: {e}")
        
        def get_pdf_content(self, pdf_path):
            with open(pdf_path, 'rb') as file:
                pdf_reader = PdfReader(file)
                pdf_content = ''
                for page_num in range(len(pdf_reader.pages)):
                    pdf_content += pdf_reader.pages[page_num].extract_text()
            return pdf_content

        def parse_json_to_dictionary(self,jsn):
             return json.loads(jsn)
